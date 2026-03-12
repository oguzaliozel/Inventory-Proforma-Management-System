import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import openpyxl
import datetime
import os
import sqlite3

class db_manager:
    def __init__(self, db_name="stok_takip.db"):
        self.conn = sqlite3.connect(db_name)
        self.cursor = self.conn.cursor()
        self.tablolari_olustur()
        self.baslangic_verilerini_ekle()

    def tablolari_olustur(self):
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS urunler (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ad TEXT,
                stok INTEGER,
                fiyat_usd REAL,
                fiyat_euro REAL,
                fiyat_tl REAL
            )
        ''')
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS musteriler (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                firma_adi TEXT,
                adres TEXT,
                vergi_no TEXT
            )
        ''')
        self.conn.commit()

    def baslangic_verilerini_ekle(self):
        self.cursor.execute("SELECT COUNT(*) FROM urunler")
        if self.cursor.fetchone()[0] == 0:
            baslangic_verileri = [
                ("OTOMATİK TEK KÖŞE KAYNAK MAKİNASI", 5, 1250, 1150, 40000),
                ("OTOMATİK ÇITA KESME MAKİNASI", 8, 1000, 920, 32000),
                ("OTOMATİK ALTTAN ÇIKMA KESİM MAKİNASI", 3, 1100, 1050, 35000),
                ("OTOMATİK ORTA KAYIT ALIŞTIRMA", 10, 1000, 920, 32000),
                ("ÜÇLÜ KOLYERİ DELME MAKİNESİ", 2, 1000, 920, 32000),
            ]
            self.cursor.executemany("INSERT INTO urunler (ad, stok, fiyat_usd, fiyat_euro, fiyat_tl) VALUES (?, ?, ?, ?, ?)", baslangic_verileri)
            self.conn.commit()

    def get_urunler(self):
        self.cursor.execute("SELECT * FROM urunler")
        return [{"id": r[0], "ad": r[1], "stok": r[2], "fiyat_usd": r[3], "fiyat_euro": r[4], "fiyat_tl": r[5]} for r in self.cursor.fetchall()]

    def urun_getir(self, urun_id):
        self.cursor.execute("SELECT * FROM urunler WHERE id=?", (urun_id,))
        r = self.cursor.fetchone()
        if r:
            return {"id": r[0], "ad": r[1], "stok": r[2], "fiyat_usd": r[3], "fiyat_euro": r[4], "fiyat_tl": r[5]}
        return None

    def urun_stok_guncelle(self, urun_id, degisim):
        # Sepete ekleme(+)/çıkarma(-) işlemleri için
        self.cursor.execute("UPDATE urunler SET stok = stok + ? WHERE id = ?", (degisim, urun_id))
        self.conn.commit()

    def urun_kaydet(self, urun_id, ad, stok, f_usd, f_eur, f_tl):
        if urun_id: # Güncelle
             self.cursor.execute("UPDATE urunler SET ad=?, stok=?, fiyat_usd=?, fiyat_euro=?, fiyat_tl=? WHERE id=?", (ad, stok, f_usd, f_eur, f_tl, urun_id))
        else: # Ekle
             self.cursor.execute("INSERT INTO urunler (ad, stok, fiyat_usd, fiyat_euro, fiyat_tl) VALUES (?, ?, ?, ?, ?)", (ad, stok, f_usd, f_eur, f_tl))
        self.conn.commit()

    def urun_sil(self, urun_id):
         self.cursor.execute("DELETE FROM urunler WHERE id=?", (urun_id,))
         self.conn.commit()

    def get_musteriler(self):
        self.cursor.execute("SELECT * FROM musteriler")
        return [{"id": r[0], "firma_adi": r[1], "adres": r[2], "vergi_no": r[3]} for r in self.cursor.fetchall()]
    
    def musteri_getir(self, musteri_id):
        self.cursor.execute("SELECT * FROM musteriler WHERE id=?", (musteri_id,))
        r = self.cursor.fetchone()
        if r:
             return {"id": r[0], "firma_adi": r[1], "adres": r[2], "vergi_no": r[3]}
        return None

    def musteri_ekle(self, firma_adi, adres, vergi_no):
        self.cursor.execute("INSERT INTO musteriler (firma_adi, adres, vergi_no) VALUES (?, ?, ?)", (firma_adi, adres, vergi_no))
        self.conn.commit()
    
    def kapat(self):
        self.conn.close()

class ProformaUygulamasi:
    def __init__(self, root):
        self.db = db_manager()
        self.root = root
        self.root.title("Stok Takip ve Proforma Sistemi V3")
        self.root.geometry("1100x700")

        # Stiller
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("TButton", font=("Segoe UI", 10), padding=5)
        style.configure("TLabel", font=("Segoe UI", 10))
        style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"), background="#e1e1e1")
        style.configure("Treeview", font=("Segoe UI", 9), rowheight=25)

        # Sepet Yapısı: Ürün objelerini ve eklenme adetlerini tutar
        self.sepet = []
        self.secili_musteri_id = None

        self.arayuz_olustur()

    def arayuz_olustur(self):
        # --- Üst Bar ---
        ust_bar = tk.Frame(self.root, bg="#2c3e50", pady=10)
        ust_bar.pack(side=tk.TOP, fill=tk.X)
        tk.Label(ust_bar, text="STOK & PROFORMA SİSTEMİ", bg="#2c3e50", fg="white", font=("Segoe UI", 16, "bold")).pack(side=tk.LEFT, padx=20)
        ttk.Button(ust_bar, text="⚙️ Yönetici Paneli", command=self.yonetici_girisi).pack(side=tk.RIGHT, padx=20)

        # --- Ana Panel Bölümü ---
        ana_panel = tk.Frame(self.root)
        ana_panel.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # --- Sol Panel: Ürün Seçimi ---
        panel_sol = ttk.LabelFrame(ana_panel, text="Ürün Listesi & Stok Durumu", padding=10)
        panel_sol.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))

        # Ürün Listesi (Treeview)
        columns = ("ID", "Ürün Adı", "Stok", "Fiyat($)", "Fiyat(€)", "Fiyat(₺)")
        self.tree_urunler = ttk.Treeview(panel_sol, columns=columns, show="headings", height=15)
        for col in columns:
            self.tree_urunler.heading(col, text=col)
            width = 250 if col == "Ürün Adı" else 60
            self.tree_urunler.column(col, width=width, anchor=tk.CENTER)
            
        self.tree_urunler.column("Ürün Adı", anchor=tk.W)
        self.tree_urunler.pack(fill=tk.BOTH, expand=True, pady=5)

        self.tabloyu_guncelle()

        # Ekle Butonu
        btn_frame = tk.Frame(panel_sol)
        btn_frame.pack(fill=tk.X, pady=5)
        ttk.Button(btn_frame, text="🛒 Seçili Ürünü Sepete Ekle", command=self.sepete_ekle).pack(side=tk.RIGHT)

        # --- Sağ Panel: Müşteri ve Proforma ---
        panel_sag = tk.Frame(ana_panel, width=350)
        panel_sag.pack(side=tk.RIGHT, fill=tk.Y, padx=(5, 0))

        # --- Müşteri Yönetimi ---
        frame_musteri = ttk.LabelFrame(panel_sag, text="Müşteri Bilgileri", padding=10)
        frame_musteri.pack(fill=tk.X, pady=(0, 10))

        # Müşteri Seçimi
        tk.Label(frame_musteri, text="Kayıtlı Müşteri Seç:", anchor="w").pack(fill=tk.X)
        self.combo_musteriler = ttk.Combobox(frame_musteri, state="readonly")
        self.combo_musteriler.pack(fill=tk.X, pady=(2, 5))
        self.combo_musteriler.bind("<<ComboboxSelected>>", self.musteri_secildi)
        
        self.musteri_combobox_guncelle()

        # Müşteri Formu
        self.entry_firma = self.input_yap(frame_musteri, "Firma Adı:")
        self.entry_adres = self.input_yap(frame_musteri, "Adres:")
        self.entry_vergi = self.input_yap(frame_musteri, "Vergi No:")

        ttk.Button(frame_musteri, text="💾 Yeni Müşteri Olarak Kaydet", command=self.musteri_kaydet).pack(fill=tk.X, pady=(10,0))

        # Alt Kısım - Proforma Butonu (Görünmez olmaması için önce pack ediyoruz)
        btn_proforma = tk.Button(panel_sag, text="📄 PROFORMA OLUŞTUR", command=self.excel_olustur, bg="#27ae60", fg="white", font=("Segoe UI", 12, "bold"), pady=10)
        btn_proforma.pack(side=tk.BOTTOM, fill=tk.X, pady=(10,0))

        # --- Sepet Paneli ---
        frame_sepet = ttk.LabelFrame(panel_sag, text="Sepet Detayları", padding=10)
        frame_sepet.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        tk.Label(frame_sepet, text="Para Birimi", anchor="w").pack(fill=tk.X)
        self.var_para_birimi = tk.StringVar(value="USD")
        cerceve_para = tk.Frame(frame_sepet)
        cerceve_para.pack(fill=tk.X, pady=(0,10))
        ttk.Radiobutton(cerceve_para, text="USD ($)", variable=self.var_para_birimi, value="USD").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(cerceve_para, text="EURO (€)", variable=self.var_para_birimi, value="EURO").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(cerceve_para, text="TL (₺)", variable=self.var_para_birimi, value="TL").pack(side=tk.LEFT, padx=5)

        self.listbox_sepet = tk.Listbox(frame_sepet, height=12, font=("Segoe UI", 9))
        self.listbox_sepet.pack(fill=tk.BOTH, expand=True, pady=5)
        
        ttk.Button(frame_sepet, text="🗑️ Sepeti Boşalt", command=self.sepeti_bosalt).pack(fill=tk.X, pady=2)

    def input_yap(self, parent, label_text):
        tk.Label(parent, text=label_text, anchor="w").pack(fill=tk.X)
        entry = ttk.Entry(parent)
        entry.pack(fill=tk.X, pady=2)
        return entry

    def musteri_combobox_guncelle(self):
        musteriler = self.db.get_musteriler()
        self.combo_musteriler['values'] = [f"{m['id']} - {m['firma_adi']}" for m in musteriler]
        self.secili_musteri_id = None

    def musteri_secildi(self, event):
        secim = self.combo_musteriler.get()
        if not secim: return
        m_id = int(secim.split(" - ")[0])
        musteri = self.db.musteri_getir(m_id)
        if musteri:
            self.secili_musteri_id = m_id
            self.entry_firma.delete(0, tk.END)
            self.entry_firma.insert(0, musteri["firma_adi"])
            self.entry_adres.delete(0, tk.END)
            self.entry_adres.insert(0, musteri["adres"])
            self.entry_vergi.delete(0, tk.END)
            self.entry_vergi.insert(0, musteri["vergi_no"])

    def musteri_kaydet(self):
        firma = self.entry_firma.get().strip()
        adres = self.entry_adres.get().strip()
        vergi = self.entry_vergi.get().strip()
        
        if not firma:
            messagebox.showwarning("Uyarı", "Firma adı boş olamaz!")
            return
            
        self.db.musteri_ekle(firma, adres, vergi)
        self.musteri_combobox_guncelle()
        messagebox.showinfo("Başarılı", "Yeni müşteri kaydedildi.")

    def tabloyu_guncelle(self):
        for i in self.tree_urunler.get_children():
            self.tree_urunler.delete(i)
        for urun in self.db.get_urunler():
            self.tree_urunler.insert("", "end", values=(urun["id"], urun["ad"], urun["stok"], urun["fiyat_usd"], urun["fiyat_euro"], urun["fiyat_tl"]))

    def sepete_ekle(self):
        selected_item = self.tree_urunler.selection()
        if not selected_item:
            messagebox.showwarning("Uyarı", "Lütfen listeden bir ürün seçin.")
            return

        item_values = self.tree_urunler.item(selected_item, "values")
        urun_id = int(item_values[0])
        db_urun = self.db.urun_getir(urun_id)
        
        if db_urun["stok"] <= 0:
            messagebox.showerror("Hata", "Bu ürün stokta kalmadı!")
            return

        sepetteki_urun = next((item for item in self.sepet if item["id"] == urun_id), None)

        if sepetteki_urun:
            sepetteki_urun["adet"] += 1
        else:
            yeni_urun = db_urun.copy()
            yeni_urun["adet"] = 1
            self.sepet.append(yeni_urun)

        # STOKTAN DÜŞÜŞÜ DB'YE YANSIT (Geçici Rezerve)
        self.db.urun_stok_guncelle(urun_id, -1)
        
        self.tabloyu_guncelle()
        self.sepet_listesini_guncelle()

    def sepet_listesini_guncelle(self):
        self.listbox_sepet.delete(0, tk.END)
        for item in self.sepet:
            self.listbox_sepet.insert(tk.END, f"{item['adet']}x {item['ad']}")

    def sepeti_bosalt(self):
        # Sepeti boşaltırken stokları DB'ye iade et
        for item in self.sepet:
             self.db.urun_stok_guncelle(item["id"], item["adet"])
        
        self.sepet = []
        self.sepet_listesini_guncelle()
        self.tabloyu_guncelle()

    # --- YÖNETİCİ PANELİ İŞLEMLERİ ---
    def yonetici_girisi(self):
        sifre = simpledialog.askstring("Giriş", "Yönetici Şifresi (1234):", show='*')
        if sifre == "1234":
            self.yonetici_penceresi_ac()
        elif sifre is not None:
            messagebox.showerror("Hata", "Yanlış Şifre!")

    def yonetici_penceresi_ac(self):
        admin_win = tk.Toplevel(self.root)
        admin_win.title("Yönetici Paneli - Ürün Yönetimi")
        admin_win.geometry("900x550")

        # Sol taraf liste
        frame_list = ttk.LabelFrame(admin_win, text="Ürünler Veritabanı", padding=10)
        frame_list.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10, pady=10)

        cols = ("ID", "Ürün Adı", "Stok", "USD", "EUR", "TL")
        tree_admin = ttk.Treeview(frame_list, columns=cols, show="headings")
        for c in cols:
            tree_admin.heading(c, text=c)
            w = 30 if c=="ID" else 200 if c=="Ürün Adı" else 60
            tree_admin.column(c, width=w, anchor="center")
        tree_admin.column("Ürün Adı", anchor="w")
        tree_admin.pack(fill=tk.BOTH, expand=True)

        # Sağ taraf form
        frame_form = ttk.LabelFrame(admin_win, text="Ürün Düzenle / Ekle", padding=10, width=300)
        frame_form.pack(side=tk.RIGHT, fill=tk.Y, padx=10, pady=10)

        entries = {}
        fields = [("Ürün Adı", "ad"), ("Stok", "stok"), ("Fiyat USD", "fiyat_usd"), ("Fiyat EURO", "fiyat_euro"), ("Fiyat TL", "fiyat_tl")]
        
        for lbl, key in fields:
            tk.Label(frame_form, text=lbl, anchor="w").pack(fill=tk.X)
            ent = ttk.Entry(frame_form)
            ent.pack(fill=tk.X, pady=(0, 5))
            entries[key] = ent

        aktif_id_var = tk.StringVar(value="")
        lbl_aktif = tk.Label(frame_form, text="Yeni Ürün Eklenecek", fg="blue", font=("", 10, "italic"))
        lbl_aktif.pack(pady=5)

        # Fonksiyonlar
        def admin_tablo_guncelle():
            for i in tree_admin.get_children(): tree_admin.delete(i)
            for urun in self.db.get_urunler():
                tree_admin.insert("", "end", values=(urun["id"], urun["ad"], urun["stok"], urun["fiyat_usd"], urun["fiyat_euro"], urun["fiyat_tl"]))

        def form_temizle():
            for ent in entries.values():
                ent.delete(0, tk.END)
            aktif_id_var.set("")
            lbl_aktif.config(text="Yeni Ürün Eklenecek", fg="blue")
            tree_admin.selection_remove(tree_admin.selection())

        def form_doldur(event):
            selected = tree_admin.selection()
            if not selected: return
            item_vals = tree_admin.item(selected, "values")
            urun_id = int(item_vals[0])
            urun = self.db.urun_getir(urun_id)
            
            form_temizle()
            aktif_id_var.set(str(urun_id))
            lbl_aktif.config(text=f"ID: {urun_id} Düzenleniyor", fg="red")
            
            for key, ent in entries.items():
                ent.insert(0, str(urun[key]))

        tree_admin.bind("<<TreeviewSelect>>", form_doldur)

        def kaydet_fonk():
            try:
                ad = entries["ad"].get().strip()
                stok = int(entries["stok"].get())
                f_usd = float(entries["fiyat_usd"].get())
                f_eur = float(entries["fiyat_euro"].get())
                f_tl = float(entries["fiyat_tl"].get())
                
                if not ad: raise ValueError("Ad boş olamaz")

                uid_str = aktif_id_var.get()
                uid = int(uid_str) if uid_str else None
                
                self.db.urun_kaydet(uid, ad, stok, f_usd, f_eur, f_tl)
                messagebox.showinfo("Başarılı", "İşlem kaydedildi!")
                admin_tablo_guncelle()
                self.tabloyu_guncelle()
                form_temizle()
            except ValueError as e:
                messagebox.showerror("Hata", f"Lütfen değerleri doğru giriniz.\n({e})")

        def sil_fonk():
            uid_str = aktif_id_var.get()
            if not uid_str:
                messagebox.showwarning("Uyarı", "Lütfen silmek için tablodan bir ürün seçin.")
                return
            cevap = messagebox.askyesno("Onay", "Bu ürünü silmek istediğinize emin misiniz?")
            if cevap:
                self.db.urun_sil(int(uid_str))
                admin_tablo_guncelle()
                self.tabloyu_guncelle()
                form_temizle()
                messagebox.showinfo("Başarılı", "Ürün silindi.")

        btn_cerceve = tk.Frame(frame_form)
        btn_cerceve.pack(fill=tk.X, pady=20)
        
        ttk.Button(btn_cerceve, text="💾 KAYDET", command=kaydet_fonk).pack(fill=tk.X, pady=2)
        ttk.Button(btn_cerceve, text="➕ Yeni Temiz Form", command=form_temizle).pack(fill=tk.X, pady=2)
        tk.Button(btn_cerceve, text="🗑️ Seçilini Sil", command=sil_fonk, bg="#e74c3c", fg="white", pady=5).pack(fill=tk.X, pady=10)

        admin_tablo_guncelle()

    # --- PROFORMA OLUŞTURMA ---
    def excel_olustur(self):
        if not self.sepet:
            messagebox.showwarning("Hata", "Sepet boş!")
            return

        para_birimi = self.var_para_birimi.get()
        firma_adi = self.entry_firma.get().strip()
        
        if not firma_adi:
             messagebox.showwarning("Eksik Bilgi", "Lütfen firma adını giriniz.")
             return

        # Dosya seçimi (Yollar 'Proforma Taslak' klasörüne yönlendirildi ve mutlak yol kullanıldı)
        base_dir = os.path.dirname(os.path.abspath(__file__))
        sablon_klasoru = os.path.join(base_dir, "Proforma  Taslak")
        
        if para_birimi == "USD":
            sablon_dosyasi = os.path.join(sablon_klasoru, "PROFORMA FATURA USD.xlsx")
            fiyat_key = "fiyat_usd"
        elif para_birimi == "EURO":
            sablon_dosyasi = os.path.join(sablon_klasoru, "PROFORMA FATURA EURO.xlsx")
            fiyat_key = "fiyat_euro"
        else:
            sablon_dosyasi = os.path.join(sablon_klasoru, "PROFORMA FATURA TL.xlsx")
            fiyat_key = "fiyat_tl"

        if not os.path.exists(sablon_dosyasi):
            messagebox.showerror("Hata", f"Şablon bulunamadı:\n{sablon_dosyasi}\nLütfen 'Proforma  Taslak' klasörünün uygulamanın yanında olduğundan emin olun.")
            return

        try:
            wb = openpyxl.load_workbook(sablon_dosyasi)
            ws = wb.active

            # --- HEADER BİLGİLERİ ---
            # B10-B13 arası birleşik (merged) hücre olduğu için bilgileri alt alta tek hücreye yazıyoruz.
            musteri_bilgisi = firma_adi
            adres_bilgisi = self.entry_adres.get().strip()
            vergi_bilgisi = self.entry_vergi.get().strip()
            
            if adres_bilgisi:
                musteri_bilgisi += f"\n{adres_bilgisi}"
            if vergi_bilgisi:
                musteri_bilgisi += f"\nVergi No: {vergi_bilgisi}"
                
            ws['B10'] = musteri_bilgisi
            ws['F8'] = datetime.datetime.now().strftime("%d.%m.%Y")

            # --- ÜRÜNLERİ YAZDIRMA ---
            # Kullanıcının talebi: İlk sıradaki "Üçlü Kolyeri..." makinesi şablonda sabit kalsın
            # Bizim sepete eklediklerimiz bu sabit makinenin altına eklensin.
            # Şablonda ilk makine 17. satırda bulunuyor. Biz eklemeye 18'den başlıyoruz.
            baslangic_satiri = 18
            
            for i, urun in enumerate(self.sepet):
                satir = baslangic_satiri + i
                fiyat = urun[fiyat_key]
                adet = urun["adet"]
                tutar = fiyat * adet

                ws.cell(row=satir, column=2).value = urun["ad"]
                ws.cell(row=satir, column=3).value = adet
                ws.cell(row=satir, column=4).value = fiyat
                # Tutar formüllü olabilir ama doğrudan değeri yazmak güvenlidir
                ws.cell(row=satir, column=5).value = tutar

            # Şablondaki geri kalan varsayılan (dummy) fiyatları temizle
            # Kendi yazdığımız son satırdan sonrasını temizliyoruz.
            # Şablonda 30, 31 ve 32. satırlar birleştirilmiş hücre (TOPLAM, KDV vs) olduğu için hata veriyordu.
            # Bu yüzden 30. satıra kadar temizlik yapıyoruz.
            son_dolu_satir = baslangic_satiri + len(self.sepet)
            for bos_satir in range(son_dolu_satir, 30):
                ws.cell(row=bos_satir, column=2).value = None
                ws.cell(row=bos_satir, column=3).value = None
                ws.cell(row=bos_satir, column=4).value = None
                ws.cell(row=bos_satir, column=5).value = None

            # Çıktılar klasörünü oluştur (yoksa)
            cikti_klasoru = os.path.join(base_dir, "Çıktılar")
            if not os.path.exists(cikti_klasoru):
                os.makedirs(cikti_klasoru)

            # Kaydet (Yeni klasöre yönlendir)
            gecerli_zaman = datetime.datetime.now().strftime("%Y%m%d_%H%M")
            yeni_dosya_adi = f"Proforma_{firma_adi[:10]}_{para_birimi}_{gecerli_zaman}.xlsx"
            
            # Geçersiz dosya karakterlerini temizle
            yeni_dosya_adi = "".join([c for c in yeni_dosya_adi if c.isalpha() or c.isdigit() or c in(' ', '.', '_')]).rstrip()
            tam_cikti_yolu = os.path.join(cikti_klasoru, yeni_dosya_adi)
            
            wb.save(tam_cikti_yolu)
            messagebox.showinfo("Başarılı", f"Proforma Oluşturuldu:\n{tam_cikti_yolu}")
            
            # Başarılı olunca sepeti kalıcı temizle (stoklar zaten DB'den düşürüldü)
            self.sepet = []
            self.sepet_listesini_guncelle()

        except PermissionError:
             messagebox.showerror("Hata", "Dosya oluşturulamadı. Aynı isimde bir Excel dosyası açıksa lütfen kapatın.")
        except Exception as e:
            messagebox.showerror("Hata", f"Excel hatası: {str(e)}")

    def __del__(self):
        # Uygulama kapanırken sepeti iade et ve DB'yi kapat
        try:
             self.sepeti_bosalt()
             self.db.kapat()
        except:
             pass

if __name__ == "__main__":
    root = tk.Tk()
    app = ProformaUygulamasi(root)
    # X ile kapatıldığında temizliği yapabilmek için (opsiyonel garanti)
    def on_closing():
        try:
            app.sepeti_bosalt()
            app.db.kapat()
        except: pass
        root.destroy()
    root.protocol("WM_DELETE_WINDOW", on_closing)
    root.mainloop()

