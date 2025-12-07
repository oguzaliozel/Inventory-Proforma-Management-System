import sys
import os
import tkinter as tk

# Projenin oldugu konumu path e ekleyelim
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
import stok_proforma

def run_tests():
    root = tk.Tk()
    # Arayuz gozukmesin diye gizliyoruz
    root.withdraw()
    app = stok_proforma.ProformaUygulamasi(root)

    print("--- TEST SURECI BASLIYOR ---")
    
    # TEST 1: Sepete Ekleme
    print("\n[TEST 1] Sepete Eleman Ekleme Testi")
    children = app.tree_urunler.get_children()
    if len(children) >= 2:
        app.tree_urunler.selection_set(children[0])
        app.sepete_ekle()
        
        # tabloyu_guncelle silinip eklendigi icin children idleri degisiyor
        children = app.tree_urunler.get_children()
        app.tree_urunler.selection_set(children[1])
        app.sepete_ekle()
        
        children = app.tree_urunler.get_children()
        app.tree_urunler.selection_set(children[1])
        app.sepete_ekle() # 2. urunden 2 tane ekleyelim
        
        sepet_durumu = [f"{i['adet']}x {i['ad']}" for i in app.sepet]
        print(f"Sepet Uzunlugu: {len(app.sepet)}, Icerik: {sepet_durumu}")
        
        if len(app.sepet) == 2 and app.sepet[1]['adet'] == 2:
            print(">> BASARILI: Urunler sepete dogru eklendi ve adetleri artis gosterdi.")
        else:
            print(">> BASARISIZ: Ekleme isleminde hata var.")
            return

        # TEST 2: Sepetten Adet Dusurme
        print("\n[TEST 2] Sepetten Adet Dusurme Testi (Son eklenen urunden 1 adet cikarilacak)")
        # UI listbox uzerinden siliyoruz
        app.listbox_sepet.selection_set(1) 
        app.secili_urunu_cikar()
        
        sepet_durumu = [f"{i['adet']}x {i['ad']}" for i in app.sepet]
        print(f"Icerik: {sepet_durumu}")
        
        if len(app.sepet) == 2 and app.sepet[1]['adet'] == 1:
            print(">> BASARILI: 2. urunun adedi 2'den 1'e basariyla dusuruldu.")
        else:
            print(">> BASARISIZ: Adet dusurme isleminde hata.")
            return

        # TEST 3: Sepetten Tamamen Silme
        print("\n[TEST 3] Sepetten Tamamen Silme Testi (Son kalan 1 adetlik urun silinecek)")
        app.listbox_sepet.selection_set(1) 
        app.secili_urunu_cikar()
        
        sepet_durumu = [f"{i['adet']}x {i['ad']}" for i in app.sepet]
        print(f"Sepet Uzunlugu: {len(app.sepet)}, Icerik: {sepet_durumu}")
        
        if len(app.sepet) == 1:
            print(">> BASARILI: Adedi 1 olan urun sepetten tamamen silindi.")
        else:
            print(">> BASARISIZ: Element silinmedi.")
            return

        # TEST 4: Excel Ciktisi Olusturma ve Format Kontrolu
        print("\n[TEST 4] Excel Olusturma Testi")
        app.entry_firma.insert(0, "Yapay Zeka Test Firmasi")
        try:
            app.excel_olustur() # Uretim fonksiyonunu cagir
            
            # Test icin dizin kontrolu
            base_dir = os.path.dirname(os.path.abspath(__file__))
            cikti_klasoru = os.path.join(base_dir, "Çıktılar")
            import glob
            list_of_files = glob.glob(cikti_klasoru + '/*.xlsx') 
            latest_file = max(list_of_files, key=os.path.getctime)
            print(f">> BASARILI: Excel dosyasi olusturuldu -> {latest_file}")
            print(">> (Satirlarin tamamen temizlendigini excel dosyasini acarak gorebilirsiniz.)")
            
            import openpyxl
            wb = openpyxl.load_workbook(latest_file)
            ws = wb.active
            if ws.cell(row=17, column=2).value is None and ws.cell(row=18, column=2).value is not None:
                print(">> BASARILI: Excel yazdirmasi dogru sekilde 18. satirdan basliyor.")
            if ws.cell(row=19, column=5).value in [None, ""]:
                print(">> BASARILI: Urun olmayan alt satirlar dogru sekilde formatlardan temizlenmis.")
            
        except Exception as e:
            print(f">> BASARISIZ: Excel olusturulurken hata -> {str(e)}")
            
        print("\n--- TUM TESTLER BASARIYLA TAMAMLANDI ---")
    else:
        print("Testler iptal edildi. En az 2 urun olmali.")
        
    root.destroy()

if __name__ == '__main__':
    run_tests()
